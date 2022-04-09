'''
*Copyright (c) 2022 All rights reserved
*@description: record the frequent itemsets and association rules in excel
*@author: Zhixing Lu
*@date: 2022-04-10
*@email: luzhixing12345@163.com
*@Github: luzhixing12345
'''

from itertools import chain
import openpyxl
from openpyxl.utils import get_column_letter

def record(frequent_itemsets, association_rules,dataset):
    '''
    *@description: record the frequent itemsets and association rules in excel
    *@param: frequent_itemsets: frequent itemsets
    *@param: association_rules: association rules
    *@return: None
    '''
    # create a workbook
    workbook = openpyxl.Workbook()
    for k,v in frequent_itemsets.items():
        sheet_name = 'frequent_itemsets' + str(k)
        sheet = workbook.create_sheet(sheet_name)
        
        start_col = 2
        start_row = 2
        sheet[get_column_letter(start_col)+str(start_row)] = 'itemsets'
        sheet[get_column_letter(start_col+1)+str(start_row)] = 'support'
        
        for items, support in v.items():
            start_row+=1
            set = "{"
            for id, item in enumerate(items):
                if id !=0:
                    set = set + ', '
                set = set + item
            set = set + '}'
            sheet[get_column_letter(start_col)+str(start_row)] = set
            sheet[get_column_letter(start_col+1)+str(start_row)] = support
            
    # write the association rules
    sheet_name = 'association_rules'
    sheet = workbook.create_sheet(sheet_name)
    start_col = 2
    start_row = 2
    sheet[get_column_letter(start_col)+str(start_row)] = 'A'
    sheet[get_column_letter(start_col+1)+str(start_row)] = 'B'
    sheet[get_column_letter(start_col+2)+str(start_row)] = 'confidence'
    
    association_rules = chain(*association_rules)
    for (A,B,confidence) in association_rules:
        start_row+=1
        set = "{"
        for id, item in enumerate(A):
            if id !=0:
                set = set + ', '
            set = set + item
        set = set + '}'
        sheet[get_column_letter(start_col)+str(start_row)] = set
        sheet[get_column_letter(start_col+1)+str(start_row)] = '{'+calculate_difference_set(A,B)+'}'
        sheet[get_column_letter(start_col+2)+str(start_row)] = confidence
        
    # save the workbook
    workbook.remove(workbook['Sheet'])
    name = 'dataset.xlsx' if dataset else 'example.xlsx'
    workbook.save(name)
    print('The result is saved in', name)
    # close the workbook
    workbook.close()


def calculate_difference_set(A,B):
    '''
    *@description: calculate the difference set of A and B
    *@param: A: set A
    *@param: B: set B
    *@return: difference_set
    '''
    for item in B:
        if item not in A:
            return item